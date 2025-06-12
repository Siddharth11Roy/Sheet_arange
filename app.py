# import streamlit as st
# import pandas as pd
# from io import BytesIO

# st.title("📊 Excel Sheet Sorter (Alphabetical Sheet Order)")

# uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

# if uploaded_file:
#     try:
#         # Read all sheets
#         all_sheets = pd.read_excel(uploaded_file, sheet_name=None, engine='openpyxl')

#         # Sort sheet names alphabetically
#         sorted_sheet_names = sorted(all_sheets.keys())

#         # Write to a new Excel file in memory
#         output = BytesIO()
#         with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
#             for sheet_name in sorted_sheet_names:
#                 all_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

#         st.success("Sheets sorted successfully!")

#         # Download button
#         st.download_button(
#             label="📥 Download Sorted Excel File",
#             data=output.getvalue(),
#             file_name="sorted_excel_sheets.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

#     except Exception as e:
#         st.error(f"❌ An error occurred: {e}")


import streamlit as st
from openpyxl import load_workbook
from io import BytesIO

st.title("📊 Excel Sheet Sorter (Preserve Formatting)")

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file:
    try:
        # Load workbook from uploaded file
        wb = load_workbook(uploaded_file)

        # Get sorted sheet names
        sorted_names = sorted(wb.sheetnames)

        # Reorder sheets
        wb._sheets = [wb[sheet_name] for sheet_name in sorted_names]

        # Save to in-memory buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Sheets reordered successfully! Formatting is preserved.")

        # Download button
        st.download_button(
            label="📥 Download Sorted Excel File",
            data=output,
            file_name="sorted_preserved_excel.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ An error occurred: {e}")

